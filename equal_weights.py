import numpy as np
import pandas as pd
import openpyxl, math, requests
from secrets1 import IEX_CLOUD_API_TOKEN

warning_message = "Whatever you entered as portfolio size was not a number!!! Please enter a valid number!"

def chunks(list_name, size):
    for i in range(0, len(list_name), size):
        yield list_name[i:i + size]


def specific_requests(symbol, rfpath):
    stocks = pd.read_csv(rfpath)
    api_url = f'https://sandbox.iexapis.com/stable/stock/{symbol}/quote/?token={IEX_CLOUD_API_TOKEN}'
    data = requests.get(api_url).json()
    price = data['latestPrice']
    market_cap = data['marketCap']
    final_dataframe.loc[len(final_dataframe)] = [symbol, price, market_cap, 'N/A']


def individual_requests(request_size, rfpath):
    stocks = pd.read_csv(rfpath)
    for stock in stocks['Ticker'][:request_size]:
        api_url = f'https://sandbox.iexapis.com/stable/stock/{stock}/quote/?token={IEX_CLOUD_API_TOKEN}'
        data = requests.get(api_url).json()
        final_dataframe.loc[len(final_dataframe)] = [stock, data['latestPrice'], data['marketCap'], 'N/A']


def batch_requests(rfpath):
    stocks = pd.read_csv(rfpath)
    symbol_groups = list(chunks(stocks['Ticker'], 100))
    symbol_strings = []
    for i in range (0, len(symbol_groups)):
        symbol_strings.append(','.join(symbol_groups[i]))
    for symbol_string in symbol_strings:
        batch_api_url = f'https://sandbox.iexapis.com/stable/stock/market/batch?symbols={symbol_string}&types=quote&token={IEX_CLOUD_API_TOKEN}'
        data = requests.get(batch_api_url).json()
        for symbol in symbol_string.split(','):
            final_dataframe.loc[len(final_dataframe)] = [symbol, data[symbol]['quote']['latestPrice'], 
                                                        data[symbol]['quote']['marketCap'], 'N/A']


def save_excel_file(fpath, request_type, final_dataframe):
    xlw = pd.ExcelWriter(fpath)
    background_color = '#0a0a23'
    font_color = '#ffffff'
    string_format = xlw.book.add_format(
        {
            'font_color': font_color,
            'bg_color' : background_color,
            'border' : 1
    })

    dollar_format = xlw.book.add_format(
        {
            'num_format': '$0.00',
            'font_color': font_color,
            'bg_color' : background_color,
            'border' : 1
    })

    integer_format = xlw.book.add_format(
        {
            'num_format': 0,
            'font_color': font_color,
            'bg_color' : background_color,
            'border' : 1
    })
    percent_format = xlw.book.add_format(
        {
            'num_format':'0.0%',
            'font_color': font_color,
            'bg_color': background_color,
            'border': 1
        })

    column_formats = {
        '1': ['A', 'Ticker', 18, string_format],
        '2': ['B', 'Stock Price', 18, dollar_format],
        '3': ['C', 'Number of shares to Buy', 25, integer_format],
        '4': ['D', 'One-Year Price Return', 25, dollar_format],
        '5': ['E', 'One-Year Return Percentile', 25, percent_format],
        '6': ['F', 'Six-Months Price Return', 25, dollar_format],
        '7': ['G', 'Six-Mont Return Percentile', 25, percent_format],
        '8': ['H', 'Three-Month Price Return', 25, dollar_format],
        '9': ['I', 'Three-Month Return Percentile', 25, percent_format],
        '10': ['J', 'One-Month] Price Return', 25, dollar_format],
        '11': ['K', 'One-Month Return Pecentile', 25, percent_format],
        '12': ['L', 'HQM Score', 16, percent_format],
        '13': ['M', 'EV/GP Percentile', 16, percent_format],
        '14': ['N', 'RV Score', 16, percent_format]
    }

    if request_type == 'individual' or request_type=='i':
        final_dataframe.to_excel(xlw, sheet_name = "Individual Rec Trades", index = False, engine='xlsxwriter')
        for column in column_formats.values():
            xlw.sheets['Individual Rec Trades'].set_column(f'{column[0]}:{column[0]}', column[2], column[3])
    elif request_type=='batch' or request_type=='b':
        final_dataframe.to_excel(xlw, sheet_name = "Batch Rec Trades", index = False, engine='xlsxwriter')
        for column in column_formats.values():
            xlw.sheets['Batch Rec Trades'].set_column(f'{column[0]}:{column[0]}', column[2], column[3])     
    elif request_type=='specific' or request_type == 's':
        final_dataframe.to_excel(xlw, sheet_name = "Specific Rec Trades", index = False, engine='xlsxwriter')   
    else:
        final_dataframe.to_excel(xlw, sheet_name = "Momentum_Value Strategy", index = False, engine='xlsxwriter')
        for column in column_formats.values():
            xlw.sheets['Momentum_Value Strategy'].set_column(f'{column[0]}:{column[0]}', column[2], column[3])
    
    xlw.close()

if __name__ == "__main__":
    dbfpath = "e:\\Python Learning\\Quant Finance\\algorithmic-trading-python-master\\starter_files\\data.xlsx"
    to_read = openpyxl.load_workbook(dbfpath)
    ws = to_read.active
    lrow = len(list(ws.rows))
    rfpath = 'e:\\Python Learning\\Quant Finance\\algorithmic-trading-python-master\\starter_files\\sp_500_stocks.csv'
    my_columns = ['Ticker', 'Stock Price', 'Market Capitalization', 'Number of shares to Buy']
    final_dataframe = pd.DataFrame(columns = my_columns)
    final_dataframe
    ref = 1
    while (ref == 1):
        request_type = ws.cell(row = lrow, column = 5).value
        portfolio_size = ws.cell(row = lrow, column = 4).value
        try:
            val = float(portfolio_size)
            request_type = request_type.lower()

            if (request_type == 'individual' or request_type == 'i'):
                integer_input = 100
                individual_requests(integer_input, rfpath)
                position_size = val/len(final_dataframe.index)
                for i in range (0, len(final_dataframe.index)):
                    final_dataframe.loc[i, 'Number of shares to Buy'] = math.floor(position_size/final_dataframe.loc[i, 'Stock Price'])
                final_dataframe.set_index('Ticker')
            

            elif (request_type == 'batch' or request_type == 'b'):
                batch_requests(rfpath)
                position_size = val/len(final_dataframe.index)
                for i in range (0, len(final_dataframe.index)):
                    final_dataframe.loc[i, 'Number of shares to Buy'] = math.floor(position_size/final_dataframe.loc[i, 'Stock Price'])
                final_dataframe.set_index('Ticker')
            

            elif (request_type=='specific' or request_type=='s'): 
                symbol = 'AAPL'#input("Enter the symbol of the company you want to check: ")
                specific_requests(symbol, rfpath)
                position_size = val/len(final_dataframe.index)
                final_dataframe.loc[0, 'Number of shares to Buy'] = position_size/final_dataframe.iloc[0, 1]
                final_dataframe.set_index('Ticker')
        
            
            else: 
                print("Sorry, got an error!!! please enter a valid input from suggestions!")
                continue
            ref = 0
        except ValueError: 
            ref = 1
            print(warning_message)
    fpath = "e:/Python Learning/Quant Finance/algorithmic-trading-python-master/starter_files/Trades recommendation.xlsx"
    save_excel_file(fpath, request_type, final_dataframe)